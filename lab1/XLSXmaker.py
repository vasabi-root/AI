import collections
from typing import List
from openpyxl.styles import (
    PatternFill, Border, Side, 
    Alignment, Font, GradientFill, Color, colors
)
from openpyxl import Workbook

from shared import Colors, Config
from node import Node

def XLSXmake(path: List[Node], XLSXname: str) -> None:
    wb = Workbook()
    ws = wb.active
    
    if (len(path) > 0):
        # СТИЛЬ ШРИФТА
        font = Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FFFFFF'
        )    
        
        alignment=Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0
        )
        
        sideThin = Side(border_style='thin', color=Colors.DARK_GREEN_xlsSTR)
        sideThick = Side(border_style='thick', color=Colors.DARK_GREEN_xlsSTR)
        
        emptyCell = PatternFill('solid', fgColor=Colors.DARK_GREEN_xlsSTR)
        simpleCell = PatternFill('solid', fgColor=Colors.GREEN_xlsSTR)
        movedCell = PatternFill('solid', fgColor=Colors.RED_STR_xlsSTR)
        
        for col in 'BCD':
            ws.column_dimensions[col].width = Config.CELL_SIZE / 9
        
        for row in range (2, len(path)*4 + 2):
            ws.row_dimensions[row].height = Config.CELL_SIZE * 0.6
            
        m = len(path[0].state)
        
        offset_row = 2
        offset_col = 2
        for i, node in enumerate(path):
            for row in range(m):
                for col in range(m):
                    cell = ws.cell(row=row+offset_row, column=col+offset_col, value=node.state[row][col])
                    border = Border(top=sideThin, bottom=sideThin, left=sideThin, right=sideThin)
                    if (row == 0):
                        border.top = sideThick
                    if (row == m-1):
                        border.bottom = sideThick
                    if (col == 0):
                        border.left = sideThick
                    if (col == m-1):
                        border.right = sideThick
                    if (i != 0 and row == z_prev_row and col == z_prev_col):
                        cell.fill = movedCell
                    else:
                        cell.fill = simpleCell
                    cell.border = border                        
                    cell.alignment = alignment
                    cell.font = font
            ws.cell(row=node.z_row+offset_row, column=node.z_col+offset_col, value='').fill = emptyCell
            ws.merge_cells(
                start_row=offset_row, end_row=offset_row+2, 
                start_column=1, end_column=1
            )
            cell = ws.cell(row=offset_row, column=1, value=node.depth)
            cell.alignment = alignment
            
            # for i in range(offset_row, offset_row+3):
            #     cell.border = Border(top=sideThick, bottom=sideThick, left=sideThick, right=sideThick)
            z_prev_row = node.z_row
            z_prev_col = node.z_col
            offset_row += 4
    else:
        ws.column_dimensions['B'].width = 10
        ws['B1'].value = 'NO SOLUTION!'
        # сохраняем и смотрим что получилось
    wb.save(XLSXname)     
    
    
    